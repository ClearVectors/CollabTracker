from flask import render_template, request, jsonify, redirect, url_for, send_from_directory, flash, send_file, Response
from werkzeug.utils import secure_filename
from app import app, db
from models import Company, Collaboration, Opportunity, Document
from datetime import datetime
from sqlalchemy import or_, func
import os
import magic
from predictive_analytics import get_predictive_analytics
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
import pandas as pd

@app.route('/pipeline')
def pipeline():
    companies = Company.query.all()
    opportunities = Opportunity.query.all()
    return render_template('pipeline.html', companies=companies, opportunities=opportunities)

# Analytics API endpoints
@app.route('/api/analytics/revenue')
def revenue_analytics():
    try:
        stage_revenue = db.session.query(
            Opportunity.stage,
            func.sum(Opportunity.expected_revenue * Opportunity.probability / 100).label('weighted_revenue')
        ).group_by(Opportunity.stage).all()
        
        collab_revenue = db.session.query(
            Collaboration.status,
            func.sum(Collaboration.kpi_revenue).label('total_revenue')
        ).group_by(Collaboration.status).all()
        
        return jsonify({
            'stage_revenue': [{
                'stage': sr[0],
                'weighted_revenue': float(sr[1] or 0)
            } for sr in stage_revenue],
            'collab_revenue': [{
                'status': cr[0],
                'total_revenue': float(cr[1] or 0)
            } for cr in collab_revenue]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/satisfaction')
def satisfaction_analytics():
    try:
        satisfaction_scores = db.session.query(
            Company.name,
            func.avg(Collaboration.kpi_satisfaction).label('avg_satisfaction')
        ).join(Collaboration).group_by(Company.name).all()
        
        return jsonify([{
            'company': score[0],
            'satisfaction': float(score[1] or 0)
        } for score in satisfaction_scores])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/pipeline')
def pipeline_analytics():
    try:
        pipeline_stats = db.session.query(
            Opportunity.stage,
            func.count(Opportunity.id).label('count'),
            func.sum(Opportunity.expected_revenue).label('total_value')
        ).group_by(Opportunity.stage).all()
        
        return jsonify([{
            'stage': stat[0],
            'count': int(stat[1]),
            'total_value': float(stat[2] or 0)
        } for stat in pipeline_stats])
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analytics/predictions')
def predictions_analytics():
    try:
        predictions = get_predictive_analytics()
        return jsonify(predictions)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export-report')
def export_report():
    try:
        # Create workbook
        wb = Workbook()
        
        # Pipeline Revenue Sheet
        ws_pipeline = wb.active
        ws_pipeline.title = "Pipeline Revenue"
        ws_pipeline['A1'] = "Pipeline Revenue by Stage"
        ws_pipeline['A1'].font = Font(bold=True, size=14)
        ws_pipeline['A2'] = "Stage"
        ws_pipeline['B2'] = "Weighted Revenue ($)"
        
        # Get pipeline revenue data
        stage_revenue = db.session.query(
            Opportunity.stage,
            func.sum(Opportunity.expected_revenue * Opportunity.probability / 100)
        ).group_by(Opportunity.stage).all()
        
        for i, (stage, revenue) in enumerate(stage_revenue, start=3):
            ws_pipeline[f'A{i}'] = stage
            ws_pipeline[f'B{i}'] = float(revenue or 0)
            
        # Add satisfaction sheet
        ws_satisfaction = wb.create_sheet("Partner Satisfaction")
        ws_satisfaction['A1'] = "Partner Satisfaction Scores"
        ws_satisfaction['A1'].font = Font(bold=True, size=14)
        ws_satisfaction['A2'] = "Company"
        ws_satisfaction['B2'] = "Satisfaction Score"
        
        satisfaction_scores = db.session.query(
            Company.name,
            func.avg(Collaboration.kpi_satisfaction)
        ).join(Collaboration).group_by(Company.name).all()
        
        for i, (company, score) in enumerate(satisfaction_scores, start=3):
            ws_satisfaction[f'A{i}'] = company
            ws_satisfaction[f'B{i}'] = float(score or 0)
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        response = send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='collaboration_analytics_report.xlsx'
        )
        
        # Add headers
        response.headers['Cache-Control'] = 'no-cache'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to generate report. Please try again.'
        }), 500

@app.route('/export-executive-report')
def export_executive_report():
    try:
        # Gather all required data
        total_pipeline_value = db.session.query(
            func.sum(Opportunity.expected_revenue * Opportunity.probability / 100)
        ).scalar() or 0
        
        active_collaborations = Collaboration.query.filter_by(status='Active').count()
        
        # Top 5 opportunities by weighted value
        top_opportunities = Opportunity.query.order_by(
            (Opportunity.expected_revenue * Opportunity.probability / 100).desc()
        ).limit(5).all()
        
        # Generate report based on format
        if request.args.get('format') == 'pdf':
            # Generate PDF report
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
            
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                name='Heading1',
                fontSize=24,
                spaceAfter=30
            ))
            styles.add(ParagraphStyle(
                name='Heading2',
                fontSize=16,
                spaceAfter=20
            ))
            
            # Build the document
            elements = []
            
            # Title
            elements.append(Paragraph("Executive Summary Report", styles['Heading1']))
            elements.append(Spacer(1, 12))
            
            # Overview section
            elements.append(Paragraph("Executive Overview", styles['Heading2']))
            overview_data = [
                ["Total Pipeline Value:", f"${total_pipeline_value:,.2f}"],
                ["Active Collaborations:", str(active_collaborations)]
            ]
            overview_table = Table(overview_data, colWidths=[200, 200])
            overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 14),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(overview_table)
            
            # Top Opportunities section
            elements.append(Paragraph("Top Opportunities", styles['Heading2']))
            opp_data = [["Title", "Company", "Weighted Value"]]
            for opp in top_opportunities:
                weighted_value = opp.expected_revenue * opp.probability / 100
                opp_data.append([
                    opp.title,
                    opp.company.name,
                    f"${weighted_value:,.2f}"
                ])
            
            opp_table = Table(opp_data, colWidths=[200, 150, 150])
            opp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(opp_table)
            
            # Generate PDF
            doc.build(elements)
            buffer.seek(0)
            
            response = send_file(
                buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name='executive_report.pdf'
            )
        else:
            # Generate Excel report
            wb = Workbook()
            ws = wb.active
            ws.title = "Executive Summary"
            
            # Title
            ws['A1'] = "Executive Summary Report"
            ws['A1'].font = Font(bold=True, size=16)
            
            # Overview section
            row = 3
            ws[f'A{row}'] = "Total Pipeline Value"
            ws[f'B{row}'] = f"${total_pipeline_value:,.2f}"
            
            row += 1
            ws[f'A{row}'] = "Active Collaborations"
            ws[f'B{row}'] = active_collaborations
            
            # Top Opportunities section
            row += 2
            ws[f'A{row}'] = "Top Opportunities"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            
            row += 1
            headers = ["Title", "Company", "Weighted Value"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=row, column=col).value = header
                ws.cell(row=row, column=col).font = Font(bold=True)
            
            for opp in top_opportunities:
                row += 1
                weighted_value = opp.expected_revenue * opp.probability / 100
                ws.cell(row=row, column=1).value = opp.title
                ws.cell(row=row, column=2).value = opp.company.name
                ws.cell(row=row, column=3).value = f"${weighted_value:,.2f}"
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            response = send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='executive_report.xlsx'
            )
        
        # Add headers
        response.headers['Cache-Control'] = 'no-cache'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': 'Failed to generate executive report. Please try again.'
        }), 500
