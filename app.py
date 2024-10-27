import os
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_from_directory, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import DeclarativeBase
from datetime import datetime
from sqlalchemy import or_, func
import magic
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
import pandas as pd

class Base(DeclarativeBase):
    pass

db = SQLAlchemy(model_class=Base)
app = Flask(__name__)

# Configure app using environment variables
app.secret_key = os.environ.get("FLASK_SECRET_KEY") or "dev_key_fortune100"
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL")
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 300,
    "pool_pre_ping": True,
}

db.init_app(app)

with app.app_context():
    import models
    db.create_all()

@app.route('/')
def dashboard():
    companies = models.Company.query.all()
    active_collaborations = models.Collaboration.query.filter_by(status='Active').all()
    opportunities = models.Opportunity.query.order_by(models.Opportunity.probability.desc()).limit(5).all()
    return render_template('dashboard.html', 
                         companies=companies, 
                         collaborations=active_collaborations,
                         opportunities=opportunities)

@app.route('/analytics')
def analytics_dashboard():
    return render_template('analytics.html')

@app.route('/api/analytics/revenue')
def revenue_analytics():
    stage_revenue = db.session.query(
        models.Opportunity.stage,
        func.sum(models.Opportunity.expected_revenue * models.Opportunity.probability / 100).label('weighted_revenue')
    ).group_by(models.Opportunity.stage).all()
    
    collab_revenue = db.session.query(
        models.Collaboration.status,
        func.sum(models.Collaboration.kpi_revenue).label('total_revenue')
    ).group_by(models.Collaboration.status).all()
    
    return jsonify({
        'stage_revenue': [{
            'stage': sr[0],
            'weighted_revenue': float(sr[1] or 0)
        } for sr in stage_revenue],
        'collab_revenue': [{
            'status': cr[0],
            'total_revenue': float(cr[1] or 0)
        } for cr in collab_revenue]
    })

@app.route('/export-report')
def export_report():
    try:
        wb = Workbook()
        
        # Pipeline Revenue Sheet
        ws_pipeline = wb.active
        ws_pipeline.title = "Pipeline Revenue"
        ws_pipeline['A1'] = "Pipeline Revenue by Stage"
        ws_pipeline['A1'].font = Font(bold=True, size=14)
        ws_pipeline['A2'] = "Stage"
        ws_pipeline['B2'] = "Weighted Revenue ($)"
        
        stage_revenue = db.session.query(
            models.Opportunity.stage,
            func.sum(models.Opportunity.expected_revenue * models.Opportunity.probability / 100)
        ).group_by(models.Opportunity.stage).all()
        
        for i, (stage, revenue) in enumerate(stage_revenue, start=3):
            ws_pipeline[f'A{i}'] = stage
            ws_pipeline[f'B{i}'] = float(revenue or 0)
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='collaboration_analytics_report.xlsx'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/export-executive-report')
def export_executive_report():
    try:
        # Gather all required data
        total_pipeline_value = db.session.query(
            func.sum(models.Opportunity.expected_revenue * models.Opportunity.probability / 100)
        ).scalar() or 0
        
        active_collaborations = models.Collaboration.query.filter_by(status='Active').count()
        
        # Generate report based on format
        if request.args.get('format') == 'pdf':
            # Generate PDF report
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
            
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                name='Heading1',
                fontSize=24,
                spaceAfter=30
            ))
            styles.add(ParagraphStyle(
                name='Heading2',
                fontSize=16,
                spaceAfter=20
            ))
            
            # Build the document
            elements = []
            
            # Title
            elements.append(Paragraph("Executive Summary Report", styles['Heading1']))
            elements.append(Spacer(1, 12))
            
            # Overview section
            elements.append(Paragraph("Executive Overview", styles['Heading2']))
            overview_data = [
                ["Total Pipeline Value:", f"${total_pipeline_value:,.2f}"],
                ["Active Collaborations:", str(active_collaborations)]
            ]
            overview_table = Table(overview_data, colWidths=[200, 200])
            overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 14),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(overview_table)
            
            # Generate PDF
            doc.build(elements)
            buffer.seek(0)
            return send_file(
                buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name='executive_report.pdf'
            )
        else:
            # Generate Excel report
            wb = Workbook()
            ws = wb.active
            ws.title = "Executive Summary"
            
            # Add data to Excel
            ws['A1'] = "Executive Summary Report"
            ws['A1'].font = Font(bold=True, size=16)
            
            row = 3
            ws[f'A{row}'] = "Total Pipeline Value"
            ws[f'B{row}'] = f"${total_pipeline_value:,.2f}"
            
            row += 1
            ws[f'A{row}'] = "Active Collaborations"
            ws[f'B{row}'] = active_collaborations
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name='executive_report.xlsx'
            )
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/pipeline')
def pipeline():
    companies = models.Company.query.all()
    opportunities = models.Opportunity.query.all()
    return render_template('pipeline.html', companies=companies, opportunities=opportunities)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
